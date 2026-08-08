from __future__ import annotations

import importlib.util
import io
import re
import tempfile
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from parser_sidecar.errors import (
    ParseError,
    PayloadTooLargeError,
    ProfileUnavailableError,
    UnsupportedDocumentError,
    UnsupportedProfileError,
)
from parser_sidecar.models import BlockType, BoundingBox

PDF_MEDIA_TYPE = "application/pdf"
DOCX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HTML_MEDIA_TYPE = "text/html"

LIGHTWEIGHT_PROFILES = {"AUTO", "DEFAULT", "FAST", "LIGHTWEIGHT"}
SUPPORTED_PROFILES = LIGHTWEIGHT_PROFILES | {"DOCLING"}


@dataclass(slots=True)
class BlockDraft:
    type: BlockType
    text: str
    page_number: int | None = None
    heading_path: list[str] = field(default_factory=list)
    bounding_box: BoundingBox | None = None
    attributes: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ParsedContent:
    title: str
    media_type: str
    engine: str
    blocks: list[BlockDraft]
    metadata: dict[str, Any] = field(default_factory=dict)


def parse_document(
    content: bytes,
    *,
    source_name: str,
    content_type: str | None,
    parser_profile: str,
    options: dict[str, Any],
    max_archive_bytes: int,
) -> ParsedContent:
    document_format, media_type = detect_document_format(content, source_name, content_type)
    profile = parser_profile.upper()
    if profile not in SUPPORTED_PROFILES:
        supported = ", ".join(sorted(SUPPORTED_PROFILES))
        raise UnsupportedProfileError(
            f"Unsupported parser profile {parser_profile!r}; supported profiles: {supported}"
        )

    if document_format in {"docx", "xlsx"}:
        _validate_office_archive(content, max_archive_bytes)

    try:
        force_ocr = _option_bool(options, "forceOcr", False)
        if profile == "DOCLING" or (document_format == "pdf" and force_ocr):
            if document_format != "pdf":
                raise UnsupportedProfileError("The DOCLING profile currently supports PDF only")
            parsed = _parse_pdf_with_docling(content, source_name)
        elif document_format == "pdf":
            parsed = _parse_pdf_with_pymupdf(content, source_name)
        elif document_format == "docx":
            parsed = _parse_docx(content, source_name)
        elif document_format == "xlsx":
            parsed = _parse_xlsx(content, source_name, options)
        elif document_format == "markdown":
            parsed = _parse_markdown_bytes(content, source_name, options)
        elif document_format == "html":
            parsed = _parse_html(content, source_name, options)
        elif document_format == "text":
            parsed = _parse_text(content, source_name, options)
        else:
            raise UnsupportedDocumentError(f"Unsupported document format: {document_format}")
    except (
        ParseError,
        PayloadTooLargeError,
        ProfileUnavailableError,
        UnsupportedDocumentError,
        UnsupportedProfileError,
    ):
        raise
    except Exception as exc:
        raise ParseError(f"Failed to parse {source_name}") from exc

    parsed.blocks = _normalize_admonition_blocks(parsed.blocks)
    if not any(block.text.strip() for block in parsed.blocks):
        hint = "; use parserProfile DOCLING for scanned PDFs" if document_format == "pdf" else ""
        raise ParseError(f"No extractable content found in {source_name}{hint}")

    parsed.media_type = media_type
    return parsed


def detect_document_format(
    content: bytes,
    source_name: str,
    content_type: str | None,
) -> tuple[str, str]:
    normalized_type = (content_type or "").split(";", 1)[0].strip().lower()
    extension = Path(source_name).suffix.lower()

    if content.startswith(b"%PDF-"):
        return "pdf", PDF_MEDIA_TYPE

    if content.startswith(b"PK"):
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                names = set(archive.namelist())
        except zipfile.BadZipFile:
            names = set()
        if any(name.startswith("word/") for name in names):
            return "docx", DOCX_MEDIA_TYPE
        if any(name.startswith("xl/") for name in names):
            return "xlsx", XLSX_MEDIA_TYPE

    if normalized_type == PDF_MEDIA_TYPE or extension == ".pdf":
        return "pdf", PDF_MEDIA_TYPE
    if normalized_type == DOCX_MEDIA_TYPE or extension == ".docx":
        return "docx", DOCX_MEDIA_TYPE
    if normalized_type == XLSX_MEDIA_TYPE or extension == ".xlsx":
        return "xlsx", XLSX_MEDIA_TYPE
    if normalized_type in {"text/markdown", "text/x-markdown"} or extension in {".md", ".markdown"}:
        return "markdown", "text/markdown"
    if normalized_type in {HTML_MEDIA_TYPE, "application/xhtml+xml"} or extension in {
        ".html",
        ".htm",
    }:
        return "html", HTML_MEDIA_TYPE
    if normalized_type.startswith("text/") or extension in {".txt", ".text"}:
        return "text", normalized_type or "text/plain"

    raise UnsupportedDocumentError(
        f"Unsupported document type {normalized_type or 'unknown'} for {source_name}"
    )


def docling_available() -> bool:
    try:
        return importlib.util.find_spec("docling") is not None
    except (ImportError, ValueError):
        return False


def _validate_office_archive(content: bytes, max_archive_bytes: int) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            expanded_size = sum(item.file_size for item in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise ParseError("Office document is not a valid ZIP package") from exc
    if expanded_size > max_archive_bytes:
        raise PayloadTooLargeError(f"Expanded Office document exceeds {max_archive_bytes} bytes")


def _parse_pdf(content: bytes, source_name: str) -> ParsedContent:
    from pypdf import PdfReader

    try:
        reader = PdfReader(io.BytesIO(content))
        if reader.is_encrypted and reader.decrypt("") == 0:
            raise ParseError("Encrypted PDF requires a password")
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError("Invalid PDF document") from exc

    raw_metadata = reader.metadata or {}
    metadata = {
        str(key).lstrip("/"): str(value) for key, value in raw_metadata.items() if value is not None
    }
    title = _clean_inline(metadata.get("Title", "")) or _fallback_title(source_name)
    blocks: list[BlockDraft] = []
    heading_path = [title] if title else []

    for page_number, page in enumerate(reader.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception as exc:
            raise ParseError(f"Failed to extract PDF page {page_number}") from exc
        for paragraph_index, paragraph in enumerate(_split_paragraphs(page_text)):
            blocks.append(
                BlockDraft(
                    type=BlockType.PARAGRAPH,
                    text=paragraph,
                    page_number=page_number,
                    heading_path=list(heading_path),
                    attributes={
                        "engine": "pypdf",
                        "pageParagraphIndex": paragraph_index,
                    },
                )
            )

    blocks = _merge_pdf_logical_blocks(blocks)
    return ParsedContent(
        title=title,
        media_type=PDF_MEDIA_TYPE,
        engine="pypdf",
        blocks=blocks,
        metadata={
            "pageCount": len(reader.pages),
            "pdfMetadata": metadata,
        },
    )


def _parse_pdf_with_pymupdf(content: bytes, source_name: str) -> ParsedContent:
    try:
        import pymupdf
    except ImportError:
        return _parse_pdf(content, source_name)

    try:
        document = pymupdf.open(stream=content, filetype="pdf")
    except Exception as exc:
        raise ParseError("Invalid PDF document") from exc
    if document.needs_pass:
        document.close()
        raise ParseError("Encrypted PDF requires a password")

    metadata = {key: value for key, value in (document.metadata or {}).items() if value}
    title = _clean_inline(metadata.get("title", "")) or _fallback_title(source_name)
    candidates: list[dict[str, Any]] = []
    font_sizes: list[float] = []
    try:
        for page_number, page in enumerate(document, start=1):
            page_dict = page.get_text("dict", sort=True)
            page_candidates: list[dict[str, Any]] = []
            for page_block_index, raw_block in enumerate(page_dict.get("blocks", [])):
                if raw_block.get("type") != 0:
                    continue
                lines: list[str] = []
                sizes: list[float] = []
                fonts: list[str] = []
                bold_characters = 0
                visible_characters = 0
                for line in raw_block.get("lines", []):
                    spans = line.get("spans", [])
                    line_text = _clean_inline("".join(str(span.get("text", "")) for span in spans))
                    if line_text:
                        lines.append(line_text)
                    for span in spans:
                        size = float(span.get("size", 0) or 0)
                        if size > 0:
                            sizes.append(size)
                            font_sizes.append(size)
                        font = str(span.get("font", "")).lower()
                        if font:
                            fonts.append(font)
                        flags = int(span.get("flags", 0) or 0)
                        span_text = str(span.get("text", ""))
                        span_characters = sum(1 for value in span_text if not value.isspace())
                        visible_characters += span_characters
                        if "bold" in font or bool(flags & 16):
                            bold_characters += span_characters
                code_like = _looks_like_pdf_code(lines, sizes)
                text = _reconstruct_pdf_text(lines, preserve_lines=code_like)
                if not text:
                    continue
                x0, y0, x1, y1 = raw_block.get("bbox", (0, 0, 0, 0))
                page_candidates.append(
                    {
                        "text": text,
                        "page": page_number,
                        "pageBlockIndex": page_block_index,
                        "bbox": (float(x0), float(y0), float(x1), float(y1)),
                        "pageWidth": float(page.rect.width),
                        "pageHeight": float(page.rect.height),
                        "fontSize": max(sizes, default=0.0),
                        "minimumFontSize": min(sizes, default=0.0),
                        "bold": bold_characters > 0,
                        "boldRatio": bold_characters / max(1, visible_characters),
                        "fonts": sorted(set(fonts)),
                        "codeLike": code_like,
                    }
                )
            candidates.extend(_order_pdf_page_candidates(page_candidates, float(page.rect.width)))
    finally:
        page_count = document.page_count
        document.close()

    candidates = _merge_pdf_admonition_candidates(candidates)
    body_size = _median(font_sizes) or 10.0
    repeated_margins = _repeated_pdf_margin_text(candidates, page_count)
    root_title = title
    heading_stack: list[str] = []
    blocks: list[BlockDraft] = []
    saw_title = False
    for candidate in candidates:
        text = candidate["text"]
        feature_gate_text = _feature_gate_list_text(text)
        if feature_gate_text is not None:
            text = feature_gate_text
        x0, y0, x1, y1 = candidate["bbox"]
        margin_key = (candidate["page"], candidate["pageBlockIndex"])
        admonition: tuple[str, str] | None = None
        if margin_key in repeated_margins:
            block_type = repeated_margins[margin_key]
            path = _dedupe_path([root_title, *heading_stack])
            heading_level = None
        else:
            admonition = _pdf_admonition(text)
            if admonition is not None:
                label, text = admonition
                heading_level = None
                block_type = BlockType.PARAGRAPH
                path = _dedupe_path([root_title, *heading_stack])
            elif candidate["codeLike"]:
                heading_level = None
                block_type = BlockType.CODE
                path = _dedupe_path([root_title, *heading_stack])
            elif feature_gate_text is not None:
                heading_level = None
                block_type = BlockType.LIST
                path = _dedupe_path([root_title, *heading_stack])
            elif _is_list_text(text):
                heading_level = None
                block_type = BlockType.LIST
                path = _dedupe_path([root_title, *heading_stack])
            else:
                heading_level = _pdf_heading_level(candidate, body_size)
            matches_document_title = (
                not saw_title
                and candidate["page"] == 1
                and _clean_inline(text).casefold() == _clean_inline(title).casefold()
            )
            if (
                admonition is None
                and not candidate["codeLike"]
                and feature_gate_text is None
                and not _is_list_text(text)
                and matches_document_title
            ):
                root_title = text
                title = text
                heading_stack.clear()
                block_type = BlockType.TITLE
                path = [text]
                saw_title = True
                heading_level = 1
            elif (
                admonition is None
                and not candidate["codeLike"]
                and feature_gate_text is None
                and not _is_list_text(text)
                and heading_level is not None
            ):
                heading_stack = _updated_heading_stack(heading_stack, heading_level, text)
                block_type = BlockType.HEADING
                path = _dedupe_path([root_title, *heading_stack])
            elif (
                admonition is None
                and not candidate["codeLike"]
                and feature_gate_text is None
                and not _is_list_text(text)
            ):
                block_type = BlockType.PARAGRAPH
                path = _dedupe_path([root_title, *heading_stack])
        attributes = {
            "engine": "pymupdf",
            "pageBlockIndex": candidate["pageBlockIndex"],
            "fontSize": round(candidate["fontSize"], 2),
            "minimumFontSize": round(candidate["minimumFontSize"], 2),
            "bodyFontSize": round(body_size, 2),
            "bold": candidate["bold"],
            "boldRatio": round(candidate["boldRatio"], 4),
            "fonts": candidate["fonts"],
            "pageWidth": candidate["pageWidth"],
            "pageHeight": candidate["pageHeight"],
            "coordinateSpace": "PDF_POINTS_TOP_LEFT",
        }
        if heading_level is not None:
            attributes["headingLevel"] = heading_level
        if candidate["codeLike"]:
            attributes["language"] = _pdf_code_language(text)
            attributes["structureDetected"] = "CODE"
        if feature_gate_text is not None:
            attributes["structureDetected"] = "FEATURE_GATE_LIST"
            attributes["itemCount"] = _feature_gate_item_count(text)
        if candidate.get("readingOrder"):
            attributes["readingOrder"] = candidate["readingOrder"]
            attributes["readingOrderIndex"] = candidate["readingOrderIndex"]
        if admonition is not None:
            attributes["admonition"] = label
            attributes["structureDetected"] = "ADMONITION"
        blocks.append(
            BlockDraft(
                type=block_type,
                text=text,
                page_number=candidate["page"],
                heading_path=path,
                bounding_box=BoundingBox(
                    x=x0, y=y0, width=max(0.0, x1 - x0), height=max(0.0, y1 - y0)
                ),
                attributes=attributes,
            )
        )

    blocks = _merge_pdf_logical_blocks(blocks)
    return ParsedContent(
        title=title,
        media_type=PDF_MEDIA_TYPE,
        engine="pymupdf",
        blocks=blocks,
        metadata={
            "pageCount": page_count,
            "pdfMetadata": metadata,
            "layoutAware": True,
            "twoColumnPageCount": len(
                {
                    candidate["page"]
                    for candidate in candidates
                    if candidate.get("readingOrder") == "TWO_COLUMN"
                }
            ),
        },
    )


def _order_pdf_page_candidates(
    candidates: list[dict[str, Any]],
    page_width: float,
) -> list[dict[str, Any]]:
    """Linearize clear two-column layouts while preserving spanning blocks.

    PyMuPDF's geometric sort is top-to-bottom, which interleaves left and right
    columns. The detector deliberately requires a wide gutter, multiple blocks
    per side, and overlapping vertical ranges so ordinary indentation is not
    mistaken for a column layout.
    """
    visual_order = sorted(
        candidates,
        key=lambda value: (
            float(value["bbox"][1]),
            float(value["bbox"][0]),
            int(value.get("pageBlockIndex", 0)),
        ),
    )
    if len(visual_order) < 4 or page_width <= 0:
        return visual_order

    narrow = [
        value
        for value in visual_order
        if float(value["bbox"][2]) - float(value["bbox"][0]) <= page_width * 0.62
    ]
    if len(narrow) < 4:
        return visual_order

    centers = sorted(
        [
            (
                (float(value["bbox"][0]) + float(value["bbox"][2])) / 2.0,
                value,
            )
            for value in narrow
        ],
        key=lambda item: (item[0], int(item[1].get("pageBlockIndex", 0))),
    )
    gaps = [(centers[index + 1][0] - centers[index][0], index) for index in range(len(centers) - 1)]
    largest_gap, split_index = max(gaps, default=(0.0, -1))
    if largest_gap < page_width * 0.16 or split_index < 1 or len(centers) - split_index - 1 < 2:
        return visual_order

    split_x = (centers[split_index][0] + centers[split_index + 1][0]) / 2.0
    left = [value for center, value in centers if center < split_x]
    right = [value for center, value in centers if center >= split_x]
    if not _pdf_columns_overlap(left, right):
        return visual_order

    left_right_edge = max(float(value["bbox"][2]) for value in left)
    right_left_edge = min(float(value["bbox"][0]) for value in right)
    if right_left_edge - left_right_edge < page_width * 0.04:
        return visual_order

    column_ids = {id(value) for value in [*left, *right]}
    spanning = [value for value in visual_order if id(value) not in column_ids]
    remaining = [*left, *right]
    ordered: list[dict[str, Any]] = []
    tolerance = 2.0

    for bridge in spanning:
        bridge_top = float(bridge["bbox"][1])
        before = [value for value in remaining if float(value["bbox"][3]) <= bridge_top + tolerance]
        ordered.extend(_pdf_column_band_order(before, split_x))
        before_ids = {id(value) for value in before}
        remaining = [value for value in remaining if id(value) not in before_ids]
        ordered.append(bridge)
    ordered.extend(_pdf_column_band_order(remaining, split_x))
    for index, candidate in enumerate(ordered):
        candidate["readingOrder"] = "TWO_COLUMN"
        candidate["readingOrderIndex"] = index
    return ordered


def _pdf_columns_overlap(
    left: list[dict[str, Any]],
    right: list[dict[str, Any]],
) -> bool:
    left_top = min(float(value["bbox"][1]) for value in left)
    left_bottom = max(float(value["bbox"][3]) for value in left)
    right_top = min(float(value["bbox"][1]) for value in right)
    right_bottom = max(float(value["bbox"][3]) for value in right)
    overlap = max(0.0, min(left_bottom, right_bottom) - max(left_top, right_top))
    shorter_span = min(left_bottom - left_top, right_bottom - right_top)
    return shorter_span > 0 and overlap / shorter_span >= 0.35


def _pdf_column_band_order(
    candidates: list[dict[str, Any]],
    split_x: float,
) -> list[dict[str, Any]]:
    def key(value: dict[str, Any]) -> tuple[float, float, int]:
        return (
            float(value["bbox"][1]),
            float(value["bbox"][0]),
            int(value.get("pageBlockIndex", 0)),
        )

    left = [
        value
        for value in candidates
        if (float(value["bbox"][0]) + float(value["bbox"][2])) / 2.0 < split_x
    ]
    left_ids = {id(value) for value in left}
    right = [value for value in candidates if id(value) not in left_ids]
    return [*sorted(left, key=key), *sorted(right, key=key)]


def _parse_pdf_with_docling(content: bytes, source_name: str) -> ParsedContent:
    if not docling_available():
        raise ProfileUnavailableError(
            "The DOCLING profile is not installed; install the 'docling' extra"
        )

    try:
        from docling.document_converter import DocumentConverter
    except ImportError as exc:
        raise ProfileUnavailableError(
            "The DOCLING profile is not installed; install the 'docling' extra"
        ) from exc

    suffix = Path(source_name).suffix or ".pdf"
    with tempfile.NamedTemporaryFile(suffix=suffix) as source_file:
        source_file.write(content)
        source_file.flush()
        try:
            result = DocumentConverter().convert(source_file.name)
            markdown = result.document.export_to_markdown()
        except Exception as exc:
            raise ParseError("Docling failed to parse the PDF") from exc

    parsed = _parse_markdown(markdown, source_name, engine="docling")
    parsed.media_type = PDF_MEDIA_TYPE
    parsed.metadata["doclingExportFormat"] = "text/markdown"
    return parsed


def _parse_docx(content: bytes, source_name: str) -> ParsedContent:
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    try:
        document = Document(io.BytesIO(content))
    except Exception as exc:
        raise ParseError("Invalid DOCX document") from exc

    items: list[Paragraph | Table] = []
    for child in document.element.body.iterchildren():
        if child.tag.endswith("}p"):
            items.append(Paragraph(child, document))
        elif child.tag.endswith("}tbl"):
            items.append(Table(child, document))

    styled_title = next(
        (
            _clean_text(item.text)
            for item in items
            if isinstance(item, Paragraph) and _heading_level(item) == 0 and _clean_text(item.text)
        ),
        "",
    )
    core_title = _clean_inline(document.core_properties.title or "")
    title = styled_title or core_title or _fallback_title(source_name)
    root_title = title
    heading_stack: list[str] = []
    blocks: list[BlockDraft] = []
    table_index = 0
    paragraph_index = 0

    for item in items:
        if isinstance(item, Paragraph):
            text = _clean_text(item.text)
            if not text:
                continue
            level = _heading_level(item)
            if level == 0:
                root_title = text
                heading_stack.clear()
                block_type = BlockType.TITLE
                heading_path = [text]
            elif level is not None:
                heading_stack = _updated_heading_stack(heading_stack, level, text)
                block_type = BlockType.HEADING
                heading_path = _dedupe_path([root_title, *heading_stack])
            elif _is_code_paragraph(item):
                block_type = BlockType.CODE
                heading_path = _dedupe_path([root_title, *heading_stack])
            elif (feature_gate_text := _feature_gate_list_text(text)) is not None:
                text = feature_gate_text
                block_type = BlockType.LIST
                heading_path = _dedupe_path([root_title, *heading_stack])
            else:
                block_type = BlockType.LIST if _is_list_paragraph(item) else BlockType.PARAGRAPH
                heading_path = _dedupe_path([root_title, *heading_stack])

            attributes: dict[str, Any] = {
                "engine": "python-docx",
                "style": item.style.name if item.style is not None else None,
                "paragraphIndex": paragraph_index + 1,
            }
            if block_type == BlockType.LIST:
                attributes["listLevel"] = _list_level(item)
            if _feature_gate_item_count(text):
                attributes["structureDetected"] = "FEATURE_GATE_LIST"
                attributes["itemCount"] = _feature_gate_item_count(text)
            blocks.append(
                BlockDraft(
                    type=block_type,
                    text=text,
                    heading_path=heading_path,
                    attributes={
                        key: value for key, value in attributes.items() if value is not None
                    },
                )
            )
            paragraph_index += 1
            continue

        rows = [[_clean_inline(cell.text) for cell in row.cells] for row in item.rows]
        rows = [row for row in rows if any(row)]
        if not rows:
            continue
        table_text = _render_markdown_table(rows)
        column_count = max((len(row) for row in rows), default=0)
        blocks.append(
            BlockDraft(
                type=BlockType.TABLE,
                text=table_text,
                heading_path=_dedupe_path([root_title, *heading_stack]),
                attributes={
                    "engine": "python-docx",
                    "format": "text/markdown",
                    "tableIndex": table_index,
                    "rowCount": len(rows),
                    "columnCount": column_count,
                },
            )
        )
        table_index += 1

    return ParsedContent(
        title=title,
        media_type=DOCX_MEDIA_TYPE,
        engine="python-docx",
        blocks=blocks,
        metadata={
            "coreProperties": {
                key: value
                for key, value in {
                    "title": core_title or None,
                    "subject": document.core_properties.subject,
                    "author": document.core_properties.author,
                    "keywords": document.core_properties.keywords,
                }.items()
                if value
            },
            "tableCount": table_index,
        },
    )


def _parse_xlsx(
    content: bytes,
    source_name: str,
    options: dict[str, Any],
) -> ParsedContent:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    data_only = _option_bool(options, "dataOnly", True)
    include_hidden = _option_bool(options, "includeHiddenSheets", False)
    max_rows = _option_int(options, "maxRowsPerSheet", 10_000, maximum=100_000)
    max_columns = _option_int(options, "maxColumns", 256, maximum=16_384)

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=data_only)
    except Exception as exc:
        raise ParseError("Invalid XLSX document") from exc

    title = _clean_inline(workbook.properties.title or "") or _fallback_title(source_name)
    blocks: list[BlockDraft] = []
    sheet_metadata: list[dict[str, Any]] = []

    try:
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible" and not include_hidden:
                continue

            regions, truncated = _worksheet_regions(worksheet, max_rows, max_columns)
            if not regions:
                continue

            heading_path = _dedupe_path([title, worksheet.title])
            blocks.append(
                BlockDraft(
                    type=BlockType.HEADING,
                    text=worksheet.title,
                    heading_path=heading_path,
                    attributes={
                        "engine": "openpyxl",
                        "sheetName": worksheet.title,
                        "sheetState": worksheet.sheet_state,
                    },
                )
            )

            for region_index, (start_row, end_row, start_column, rows) in enumerate(regions):
                end_column = start_column + max(len(row) for row in rows) - 1
                cell_range = (
                    f"{get_column_letter(start_column)}{start_row}:"
                    f"{get_column_letter(end_column)}{end_row}"
                )
                blocks.append(
                    BlockDraft(
                        type=BlockType.TABLE,
                        text=_render_markdown_table(rows),
                        heading_path=heading_path,
                        attributes={
                            "engine": "openpyxl",
                            "format": "text/markdown",
                            "sheetName": worksheet.title,
                            "regionIndex": region_index,
                            "cellRange": cell_range,
                            "rowCount": len(rows),
                            "columnCount": max(len(row) for row in rows),
                            "truncated": truncated,
                        },
                    )
                )

            sheet_metadata.append(
                {
                    "name": worksheet.title,
                    "state": worksheet.sheet_state,
                    "regionCount": len(regions),
                    "truncated": truncated,
                }
            )
    finally:
        workbook.close()

    return ParsedContent(
        title=title,
        media_type=XLSX_MEDIA_TYPE,
        engine="openpyxl",
        blocks=blocks,
        metadata={
            "sheets": sheet_metadata,
            "dataOnly": data_only,
        },
    )


def _worksheet_regions(
    worksheet: Any,
    max_rows: int,
    max_columns: int,
) -> tuple[list[tuple[int, int, int, list[list[str]]]], bool]:
    regions: list[tuple[int, int, int, list[list[str]]]] = []
    current_rows: list[tuple[int, list[str]]] = []
    truncated = False
    worksheet_columns = worksheet.max_column or 1
    effective_max_columns = min(max_columns, max(1, worksheet_columns))

    def flush() -> None:
        if not current_rows:
            return
        non_empty_columns = [
            column_index
            for _, row in current_rows
            for column_index, value in enumerate(row)
            if value
        ]
        if not non_empty_columns:
            current_rows.clear()
            return
        first_column = min(non_empty_columns)
        last_column = max(non_empty_columns)
        trimmed_rows = [row[first_column : last_column + 1] for _, row in current_rows]
        regions.append(
            (
                current_rows[0][0],
                current_rows[-1][0],
                first_column + 1,
                trimmed_rows,
            )
        )
        current_rows.clear()

    for row_number, values in enumerate(
        worksheet.iter_rows(max_col=effective_max_columns, values_only=True),
        start=1,
    ):
        if row_number > max_rows:
            truncated = True
            break
        rendered = [_spreadsheet_value(value) for value in values]
        while rendered and not rendered[-1]:
            rendered.pop()
        if any(rendered):
            current_rows.append((row_number, rendered))
        else:
            flush()
    flush()

    return regions, truncated


def _parse_markdown_bytes(
    content: bytes,
    source_name: str,
    options: dict[str, Any],
) -> ParsedContent:
    text = _decode_text(content, options)
    return _parse_markdown(text, source_name, engine="markdown")


def _parse_markdown(text: str, source_name: str, *, engine: str) -> ParsedContent:
    lines = text.replace("\r\n", "\n").replace("\r", "\n").splitlines()
    first_title = next(
        (
            _clean_inline(_split_heading_anchor(match.group(2))[0])
            for line in lines
            if (match := re.match(r"^(#{1})\s+(.+?)\s*#*\s*$", line.strip()))
        ),
        "",
    )
    title = first_title or _fallback_title(source_name)
    root_title = title
    heading_stack: list[str] = []
    blocks: list[BlockDraft] = []
    paragraph_lines: list[str] = []
    paragraph_start: int | None = None

    def heading_path() -> list[str]:
        return _dedupe_path([root_title, *heading_stack])

    def flush_paragraph() -> None:
        nonlocal paragraph_lines, paragraph_start
        text_value = _clean_markdown_text("\n".join(paragraph_lines))
        line_start = paragraph_start
        line_end = line_start + len(paragraph_lines) - 1 if line_start is not None else None
        paragraph_lines = []
        paragraph_start = None
        if not text_value:
            return
        list_lines = [line for line in text_value.splitlines() if line.strip()]
        feature_gate_text = _feature_gate_list_text(text_value)
        if feature_gate_text is not None:
            text_value = feature_gate_text
        is_list = feature_gate_text is not None or (
            bool(list_lines) and all(_is_markdown_list_line(line) for line in list_lines)
        )
        blocks.append(
            BlockDraft(
                type=BlockType.LIST if is_list else BlockType.PARAGRAPH,
                text=text_value,
                heading_path=heading_path(),
                attributes={
                    "engine": engine,
                    **(
                        {
                            "structureDetected": "FEATURE_GATE_LIST",
                            "itemCount": _feature_gate_item_count(text_value),
                        }
                        if feature_gate_text is not None
                        else {}
                    ),
                    **(
                        {"lineStart": line_start, "lineEnd": line_end}
                        if line_start is not None
                        else {}
                    ),
                },
            )
        )

    index = 0
    while index < len(lines):
        stripped = lines[index].strip()
        heading_match = re.match(r"^(#{1,6})\s+(.+?)\s*#*\s*$", stripped)
        if heading_match:
            flush_paragraph()
            level = len(heading_match.group(1))
            heading_source, heading_anchor = _split_heading_anchor(heading_match.group(2))
            heading_text = _clean_inline(heading_source)
            if level == 1 and heading_text == title:
                root_title = heading_text
                heading_stack.clear()
                block_type = BlockType.TITLE
                path = [heading_text]
            else:
                heading_stack = _updated_heading_stack(heading_stack, level, heading_text)
                block_type = BlockType.HEADING
                path = heading_path()
            blocks.append(
                BlockDraft(
                    type=block_type,
                    text=heading_text,
                    heading_path=path,
                    attributes={
                        "engine": engine,
                        "headingLevel": level,
                        **({"anchor": heading_anchor} if heading_anchor else {}),
                        "lineStart": index + 1,
                        "lineEnd": index + 1,
                    },
                )
            )
            index += 1
            continue

        if stripped.startswith("```"):
            flush_paragraph()
            language = stripped[3:].strip()
            code_start = index + 1
            code_lines: list[str] = []
            index += 1
            while index < len(lines) and not lines[index].strip().startswith("```"):
                code_lines.append(lines[index])
                index += 1
            if index < len(lines):
                index += 1
            code = "\n".join(code_lines).strip("\n")
            if code:
                blocks.append(
                    BlockDraft(
                        type=BlockType.CODE,
                        text=code,
                        heading_path=heading_path(),
                        attributes={
                            key: value
                            for key, value in {
                                "engine": engine,
                                "language": language or None,
                                "lineStart": code_start,
                                "lineEnd": index,
                            }.items()
                            if value is not None
                        },
                    )
                )
            continue

        if _starts_markdown_table(lines, index):
            flush_paragraph()
            table_start = index + 1
            table_lines = [lines[index].strip(), lines[index + 1].strip()]
            index += 2
            while index < len(lines) and "|" in lines[index] and lines[index].strip():
                table_lines.append(lines[index].strip())
                index += 1
            blocks.append(
                BlockDraft(
                    type=BlockType.TABLE,
                    text="\n".join(table_lines),
                    heading_path=heading_path(),
                    attributes={
                        "engine": engine,
                        "format": "text/markdown",
                        "lineStart": table_start,
                        "lineEnd": index,
                    },
                )
            )
            continue

        if not stripped:
            flush_paragraph()
        else:
            if paragraph_start is None:
                paragraph_start = index + 1
            paragraph_lines.append(lines[index])
        index += 1

    flush_paragraph()
    return ParsedContent(
        title=title,
        media_type="text/markdown",
        engine=engine,
        blocks=blocks,
    )


def _parse_html(
    content: bytes,
    source_name: str,
    options: dict[str, Any],
) -> ParsedContent:
    from lxml import html

    explicit_encoding = options.get("encoding")
    decoded = _decode_text(
        content,
        {"encoding": explicit_encoding} if explicit_encoding is not None else {},
    )
    parser = html.HTMLParser(
        recover=True,
        no_network=True,
    )
    try:
        document = html.document_fromstring(decoded, parser=parser)
    except (ValueError, TypeError) as exc:
        raise ParseError(f"Unable to parse HTML in {source_name}") from exc

    for element in document.xpath("//script|//style|//noscript|//template|//nav"):
        element.drop_tree()

    title_values = document.xpath("//title[1]//text()")
    heading_values = document.xpath("//h1[1]//text()")
    title = _clean_inline(" ".join(title_values or heading_values)) or _fallback_title(source_name)
    root_title = title
    heading_stack: list[str] = []
    blocks: list[BlockDraft] = []
    saw_title_heading = False

    candidates = document.xpath(
        "//body//*[self::h1 or self::h2 or self::h3 or self::h4 or self::h5 or self::h6 "
        "or self::p or self::li or self::pre or self::table or self::figcaption "
        "or self::img[@alt]]"
    )
    for element in candidates:
        tag = str(element.tag).lower()
        if tag in {"p", "li", "figcaption", "img"} and element.xpath(
            "ancestor::table or ancestor::pre"
        ):
            continue
        if tag == "p" and element.xpath("ancestor::li"):
            continue

        if tag == "table":
            rows = [
                [_clean_inline(" ".join(cell.itertext())) for cell in row.xpath("./th|./td")]
                for row in element.xpath(".//tr")
            ]
            rows = [row for row in rows if any(row)]
            text_value = _render_markdown_table(rows)
            block_type = BlockType.TABLE
        elif tag == "img":
            text_value = _clean_inline(element.get("alt", ""))
            block_type = BlockType.CAPTION
        elif tag == "li":
            text_value = _clean_inline(
                " ".join(element.xpath("./text() | ./*[not(self::ul or self::ol)]//text()"))
            )
            block_type = BlockType.LIST
        else:
            text_value = _clean_text("\n".join(element.itertext()))
            block_type = {
                "pre": BlockType.CODE,
                "figcaption": BlockType.CAPTION,
            }.get(tag, BlockType.PARAGRAPH)

        if not text_value:
            continue
        feature_gate_text = (
            _feature_gate_list_text(text_value)
            if tag not in {"pre", "table"} and not re.fullmatch(r"h[1-6]", tag)
            else None
        )
        if feature_gate_text is not None:
            text_value = feature_gate_text
            block_type = BlockType.LIST
        if re.fullmatch(r"h[1-6]", tag):
            level = int(tag[1])
            if level == 1 and not saw_title_heading and text_value == title:
                root_title = text_value
                heading_stack.clear()
                block_type = BlockType.TITLE
                path = [root_title]
                saw_title_heading = True
            else:
                heading_stack = _updated_heading_stack(heading_stack, level, text_value)
                block_type = BlockType.HEADING
                path = _dedupe_path([root_title, *heading_stack])
            attributes = {"engine": "lxml-html", "headingLevel": level}
        else:
            path = _dedupe_path([root_title, *heading_stack])
            attributes = {
                "engine": "lxml-html",
                **({"format": "text/markdown"} if tag == "table" else {}),
                **(
                    {
                        "structureDetected": "FEATURE_GATE_LIST",
                        "itemCount": _feature_gate_item_count(text_value),
                    }
                    if feature_gate_text is not None
                    else {}
                ),
            }
        blocks.append(
            BlockDraft(
                type=block_type,
                text=text_value,
                heading_path=path,
                attributes=attributes,
            )
        )

    if not any(block.type == BlockType.TITLE for block in blocks):
        blocks.insert(
            0,
            BlockDraft(
                type=BlockType.TITLE,
                text=title,
                heading_path=[title],
                attributes={"engine": "lxml-html", "synthetic": True},
            ),
        )
    return ParsedContent(
        title=title,
        media_type=HTML_MEDIA_TYPE,
        engine="lxml-html",
        blocks=blocks,
    )


def _parse_text(
    content: bytes,
    source_name: str,
    options: dict[str, Any],
) -> ParsedContent:
    text = _decode_text(content, options)
    title = _fallback_title(source_name)
    blocks = [
        BlockDraft(
            type=BlockType.PARAGRAPH,
            text=paragraph,
            heading_path=[title] if title else [],
            attributes={"engine": "plain-text"},
        )
        for paragraph in _split_paragraphs(text)
    ]
    return ParsedContent(
        title=title,
        media_type="text/plain",
        engine="plain-text",
        blocks=blocks,
    )


def _decode_text(content: bytes, options: dict[str, Any]) -> str:
    explicit_encoding = options.get("encoding")
    if explicit_encoding is not None:
        try:
            return content.decode(str(explicit_encoding))
        except (LookupError, UnicodeDecodeError) as exc:
            raise ParseError(f"Unable to decode text with encoding {explicit_encoding!r}") from exc

    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        return content.decode("utf-16")
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ParseError("Text is not valid UTF-8; provide options.encoding") from exc


def _heading_level(paragraph: Any) -> int | None:
    style = paragraph.style
    style_name = style.name if style is not None else ""
    style_id = style.style_id if style is not None else ""
    if style_name.strip().lower() == "title" or style_id.lower() == "title":
        return 0

    for candidate in (style_name, style_id):
        match = re.search(r"(?:heading|head|标题|標題)\s*[-_]?\s*(\d+)", candidate, re.I)
        if match:
            return max(1, int(match.group(1)))

    try:
        outline_values = paragraph._p.xpath("./w:pPr/w:outlineLvl/@w:val")
    except (AttributeError, TypeError):
        outline_values = []
    if outline_values:
        return int(outline_values[0]) + 1
    return None


def _is_list_paragraph(paragraph: Any) -> bool:
    style_name = paragraph.style.name.lower() if paragraph.style is not None else ""
    paragraph_properties = paragraph._p.pPr
    return "list" in style_name or (
        paragraph_properties is not None and paragraph_properties.numPr is not None
    )


def _is_code_paragraph(paragraph: Any) -> bool:
    style = paragraph.style
    style_name = style.name if style is not None else ""
    style_id = style.style_id if style is not None else ""
    return any(
        token in candidate.strip().lower()
        for candidate in (style_name, style_id)
        for token in (
            "source code",
            "sourcecode",
            "code block",
            "codeblock",
            "preformatted",
            "代码",
        )
    )


def _list_level(paragraph: Any) -> int | None:
    paragraph_properties = paragraph._p.pPr
    if (
        paragraph_properties is None
        or paragraph_properties.numPr is None
        or paragraph_properties.numPr.ilvl is None
    ):
        return None
    return int(paragraph_properties.numPr.ilvl.val)


def _updated_heading_stack(current: list[str], level: int, text: str) -> list[str]:
    index = max(0, level - 1)
    updated = list(current[:index])
    while len(updated) < index:
        updated.append("")
    updated.append(text)
    return updated


def _dedupe_path(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        cleaned = _clean_inline(value)
        if cleaned and (not result or result[-1] != cleaned):
            result.append(cleaned)
    return result


def _split_paragraphs(text: str) -> list[str]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    parts = re.split(r"\n[ \t]*\n+", normalized)
    return [cleaned for part in parts if (cleaned := _clean_text(part))]


def _clean_text(value: str) -> str:
    lines = [re.sub(r"[ \t\f\v]+", " ", line).strip() for line in value.splitlines()]
    while lines and not lines[0]:
        lines.pop(0)
    while lines and not lines[-1]:
        lines.pop()
    return "\n".join(lines)


def _clean_markdown_text(value: str) -> str:
    lines = [line.rstrip(" \t\f\v") for line in value.splitlines()]
    while lines and not lines[0].strip():
        lines.pop(0)
    while lines and not lines[-1].strip():
        lines.pop()
    return "\n".join(lines)


def _clean_inline(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _split_heading_anchor(value: str) -> tuple[str, str | None]:
    match = re.search(r"\s*\{#([\w:.-]+)\}\s*$", value)
    if match is None:
        return value, None
    return value[: match.start()].rstrip(), match.group(1)


def _fallback_title(source_name: str) -> str:
    return Path(source_name).stem.strip() or "document"


def _render_markdown_table(rows: list[list[str]]) -> str:
    column_count = max((len(row) for row in rows), default=0)
    if column_count == 0:
        return ""
    normalized_rows = [
        [
            _escape_markdown_cell(row[index] if index < len(row) else "")
            for index in range(column_count)
        ]
        for row in rows
    ]
    header = normalized_rows[0]
    separator = ["---"] * column_count
    body = normalized_rows[1:]
    return "\n".join("| " + " | ".join(row) + " |" for row in [header, separator, *body])


def _escape_markdown_cell(value: str) -> str:
    return _clean_inline(value).replace("\\", "\\\\").replace("|", "\\|")


def _spreadsheet_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return _clean_inline(str(value))


def _option_bool(options: dict[str, Any], name: str, default: bool) -> bool:
    value = options.get(name, default)
    if isinstance(value, bool):
        return value
    if isinstance(value, str) and value.lower() in {"true", "false"}:
        return value.lower() == "true"
    raise ParseError(f"options.{name} must be a boolean")


def _option_int(
    options: dict[str, Any],
    name: str,
    default: int,
    *,
    maximum: int,
) -> int:
    value = options.get(name, default)
    if isinstance(value, bool):
        raise ParseError(f"options.{name} must be an integer")
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ParseError(f"options.{name} must be an integer") from exc
    if parsed <= 0 or parsed > maximum:
        raise ParseError(f"options.{name} must be between 1 and {maximum}")
    return parsed


def _is_markdown_list_line(line: str) -> bool:
    return re.match(r"^\s*(?:[-+*•]|\d+[.)])\s+\S", line) is not None


def _starts_markdown_table(lines: list[str], index: int) -> bool:
    if index + 1 >= len(lines) or "|" not in lines[index]:
        return False
    separator = lines[index + 1].strip().strip("|")
    cells = [cell.strip() for cell in separator.split("|")]
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells)


def _median(values: list[float]) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    middle = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[middle]
    return (ordered[middle - 1] + ordered[middle]) / 2


def _pdf_heading_level(candidate: dict[str, Any], body_size: float) -> int | None:
    text = str(candidate["text"])
    if (
        len(text) > 180
        or text.count("\n") > 2
        or _is_markdown_list_line(text)
        or _pdf_admonition(text) is not None
        or _looks_like_code_text(text)
    ):
        return None
    size = float(candidate["fontSize"])
    ratio = size / max(1.0, body_size)
    looks_terminal = bool(re.search(r"[。！？.!?；;]$", text))
    if ratio >= 1.65:
        return 1
    if ratio >= 1.35:
        return 2
    if ratio >= 1.16 or (
        float(candidate.get("boldRatio", 0.0)) >= 0.85 and not looks_terminal and len(text) <= 100
    ):
        return 3
    return None


def _reconstruct_pdf_text(lines: list[str], *, preserve_lines: bool) -> str:
    cleaned = [_clean_inline(line) for line in lines if _clean_inline(line)]
    if not cleaned:
        return ""
    if preserve_lines:
        return _reconstruct_pdf_code_lines(cleaned)

    result = cleaned[0]
    for line in cleaned[1:]:
        if result.endswith("•"):
            result += " " + line
            continue
        if re.search(r"(?:^|\n)\d+[.)]$", result):
            result += " " + line
            continue
        previous = result[-1]
        following = line[0]
        if _is_cjk(previous) or _is_cjk(following):
            separator = ""
        elif previous in "([{/<" or following in ")]},.;:!?%>/":
            separator = ""
        elif previous == "-" and following.isalpha():
            separator = ""
        else:
            separator = " "
        result += separator + line
    return result


def _reconstruct_pdf_code_lines(lines: list[str]) -> str:
    result = lines[0]
    previous = lines[0]
    for line in lines[1:]:
        separator = "" if _is_pdf_visual_code_wrap(previous, line) else "\n"
        result += separator + line
        previous = line
    return result


def _is_pdf_visual_code_wrap(previous: str, following: str) -> bool:
    left = previous.rstrip()
    right = following.lstrip()
    if not left or not right:
        return False
    if left[-1] in "\\_-." and (right[0].isalnum() or right[0] in "_-"):
        return True
    if len(left) < 72:
        return False
    if right[0] in ":,.;)]}" and left[-1] not in ";{}":
        return True
    if left[-1] == ":" and (right[0].isalnum() or right[0] in "\"'[{(-"):
        return True
    return (left[-1].islower() or left[-1].isdigit() or left[-1] == "_") and (
        right[0].islower() or right[0].isdigit() or right[0] == "_"
    )


def _is_cjk(value: str) -> bool:
    return bool(value) and (
        "\u3400" <= value <= "\u4dbf"
        or "\u4e00" <= value <= "\u9fff"
        or "\u3040" <= value <= "\u30ff"
        or "\uac00" <= value <= "\ud7af"
    )


def _looks_like_pdf_code(lines: list[str], sizes: list[float]) -> bool:
    cleaned = [_clean_inline(line) for line in lines if _clean_inline(line)]
    if not cleaned:
        return False
    text = "\n".join(cleaned)
    if text.startswith("```") or text.endswith("```"):
        return True
    signals = sum(1 for line in cleaned if _looks_like_code_text(line))
    if len(cleaned) == 1:
        return signals == 1
    return signals >= 2 and signals / len(cleaned) >= 0.4


def _looks_like_code_text(value: str) -> bool:
    text = value.strip()
    return bool(
        re.search(
            r"^(?:CREATE|DROP|SELECT|INSERT|UPDATE|DELETE|ALTER|RETURNS?|PROPERTIES|AS\s+\$\$|"
            r"WITH\b|FROM\b|WHERE\b|GROUP\s+BY\b|ORDER\s+BY\b|HAVING\b|LIMIT\b|"
            r"(?:LEFT|RIGHT|INNER|OUTER|CROSS|FULL)?\s*JOIN\b|ON\b|AND\b|OR\b|"
            r"CASE\b|WHEN\b|THEN\b|ELSE\b|END\b|"
            r"(?:mysql|psql|sqlite)\s*>|"
            r"def\s+\w+\s*\(|class\s+\w+|from\s+\S+\s+import\s+|import\s+\S+|"
            r"if\s+.+:|elif\s+.+:|else:|for\s+.+:|while\s+.+:|return\b|"
            r"[}\])]\s*(?:[A-Za-z_]\w*)?\s*;?$|"
            r"[\"']?[\w.-]+[\"']?\s*[=:]\s*[\"'\d[{]|"
            r"[A-Za-z][A-Za-z0-9_. -]{0,40}\s*:\s*\S)",
            text,
            re.IGNORECASE,
        )
        or "$$" in text
        or re.search(r"\b(?:NULL|BOOLEAN|VARCHAR|STRING|INT|BIGINT)\s*[,)]", text, re.I)
    )


def _pdf_admonition(value: str) -> tuple[str, str] | None:
    match = re.match(
        r"^:::(caution|warning|danger|tip|note|info)\s+(.*?)\s*:::$",
        value.strip(),
        re.IGNORECASE | re.DOTALL,
    )
    if match is None:
        return None
    kind = match.group(1).lower()
    label = _admonition_label(kind)
    body = match.group(2).strip()
    body = re.sub(rf"^{re.escape(label)}[：:]?\s*", "", body)
    return kind, f"{label}：{body}"


def _admonition_label(kind: str) -> str:
    return {
        "caution": "注意",
        "warning": "警告",
        "danger": "危险",
        "tip": "提示",
        "note": "说明",
        "info": "信息",
    }[kind]


def _normalize_admonition_blocks(blocks: list[BlockDraft]) -> list[BlockDraft]:
    normalized: list[BlockDraft] = []
    active_kind: str | None = None
    label_pending = False
    opener = re.compile(r"^:::(caution|warning|danger|tip|note|info)\b\s*", re.I)

    for block in blocks:
        text = block.text.strip()
        match = opener.match(text)
        if active_kind is None and match is not None:
            active_kind = match.group(1).lower()
            label_pending = True
            text = text[match.end() :].strip()

        if active_kind is None:
            normalized.append(block)
            continue

        closes = bool(re.search(r"\s*:::\s*$", text))
        if closes:
            text = re.sub(r"\s*:::\s*$", "", text).strip()
        if text:
            label = _admonition_label(active_kind)
            if label_pending:
                text = re.sub(rf"^{re.escape(label)}[：:]?\s*", "", text)
                text = f"{label}：{text}" if text else label
                label_pending = False
            normalized.append(
                BlockDraft(
                    type=block.type,
                    text=text,
                    page_number=block.page_number,
                    heading_path=block.heading_path,
                    bounding_box=block.bounding_box,
                    attributes={
                        **block.attributes,
                        "admonition": active_kind,
                        "structureDetected": "ADMONITION",
                    },
                )
            )
        if closes:
            active_kind = None
            label_pending = False

    return normalized


def _pdf_code_language(value: str) -> str:
    if re.search(r"\b(?:SELECT|CREATE|DROP|ALTER|INSERT|UPDATE|DELETE)\b", value, re.I):
        return "sql"
    if re.search(r"^(?:def|class|from|import)\b|\b(?:None|True|False)\b", value, re.M):
        return "python"
    return ""


def _merge_pdf_admonition_candidates(
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    index = 0
    while index < len(candidates):
        candidate = candidates[index]
        text = str(candidate["text"]).strip()
        if not re.match(r"^:::(?:caution|warning|danger|tip|note|info)\b", text, re.I):
            merged.append(candidate)
            index += 1
            continue
        values = [text]
        current = dict(candidate)
        cursor = index + 1
        while not values[-1].endswith(":::") and cursor < len(candidates):
            following = candidates[cursor]
            if following["page"] != candidate["page"]:
                break
            values.append(str(following["text"]).strip())
            x0 = min(float(current["bbox"][0]), float(following["bbox"][0]))
            y0 = min(float(current["bbox"][1]), float(following["bbox"][1]))
            x1 = max(float(current["bbox"][2]), float(following["bbox"][2]))
            y1 = max(float(current["bbox"][3]), float(following["bbox"][3]))
            current["bbox"] = (x0, y0, x1, y1)
            current["fontSize"] = max(float(current["fontSize"]), float(following["fontSize"]))
            current["minimumFontSize"] = min(
                float(current["minimumFontSize"]), float(following["minimumFontSize"])
            )
            current["fonts"] = sorted(set(current["fonts"]) | set(following["fonts"]))
            cursor += 1
        current["text"] = _reconstruct_pdf_text(values, preserve_lines=False)
        merged.append(current)
        index = cursor
    return merged


def _merge_pdf_logical_blocks(blocks: list[BlockDraft]) -> list[BlockDraft]:
    merged: list[BlockDraft] = []
    for block in blocks:
        if not merged:
            merged.append(block)
            continue
        previous = merged[-1]
        same_path = previous.heading_path == block.heading_path
        merge_code = (
            previous.type == block.type == BlockType.CODE
            and previous.page_number == block.page_number
            and same_path
        )
        merge_cross_page_code = _is_pdf_cross_page_code_continuation(previous, block)
        merge_paragraph = (
            previous.type == block.type == BlockType.PARAGRAPH
            and previous.page_number == block.page_number
            and same_path
            and "admonition" not in previous.attributes
            and "admonition" not in block.attributes
            and not re.search(r"[。！？.!?；;：:]$", previous.text)
        )
        if not (merge_code or merge_cross_page_code or merge_paragraph):
            merged.append(block)
            continue
        separator = (
            ""
            if merge_cross_page_code and _is_pdf_visual_code_wrap(previous.text, block.text)
            else "\n"
            if merge_code or merge_cross_page_code
            else ""
            if _is_cjk(previous.text[-1]) or _is_cjk(block.text[0])
            else " "
        )
        page_end = block.page_number or previous.page_number
        attributes = {
            **previous.attributes,
            "pageEnd": page_end,
            "mergedBlockCount": int(previous.attributes.get("mergedBlockCount", 1)) + 1,
        }
        language = previous.attributes.get("language") or block.attributes.get("language")
        if language:
            attributes["language"] = language
        merged[-1] = BlockDraft(
            type=BlockType.CODE if merge_cross_page_code else previous.type,
            text=previous.text + separator + block.text,
            page_number=previous.page_number,
            heading_path=previous.heading_path,
            bounding_box=(
                previous.bounding_box
                if previous.page_number != block.page_number
                else _union_bbox(previous.bounding_box, block.bounding_box)
            ),
            attributes=attributes,
        )
    return merged


def _is_pdf_cross_page_code_continuation(
    previous: BlockDraft,
    current: BlockDraft,
) -> bool:
    previous_page = int(previous.attributes.get("pageEnd") or previous.page_number or 0)
    current_page = int(current.page_number or 0)
    if current_page != previous_page + 1:
        return False
    if previous.heading_path != current.heading_path:
        return False
    if previous.bounding_box is None or current.bounding_box is None:
        return False

    previous_height = float(previous.attributes.get("pageHeight") or 0)
    current_height = float(current.attributes.get("pageHeight") or 0)
    if previous_height <= 0 or current_height <= 0:
        return False
    previous_bottom = previous.bounding_box.y + previous.bounding_box.height
    current_top = current.bounding_box.y
    if previous_bottom < previous_height * 0.78 or current_top > current_height * 0.22:
        return False

    starts_with_continuation = bool(
        re.match(
            r"^(?:[}\]),:]|\)\s*(?:[A-Za-z_]\w*)?(?:\s|$)|"
            r"WHERE\b|GROUP\s+BY\b|ORDER\s+BY\b|HAVING\b|LIMIT\b|"
            r"(?:LEFT|RIGHT|INNER|OUTER|CROSS|FULL)?\s*JOIN\b|ON\b|AND\b|OR\b)",
            current.text.lstrip(),
            re.IGNORECASE,
        )
    )
    if not starts_with_continuation:
        return False

    previous_code_like = previous.type == BlockType.CODE or _looks_like_pdf_code(
        previous.text.splitlines(), []
    )
    current_code_like = current.type == BlockType.CODE or _looks_like_pdf_code(
        current.text.splitlines(), []
    )
    return previous_code_like and current_code_like


def _union_bbox(left: BoundingBox | None, right: BoundingBox | None) -> BoundingBox | None:
    if left is None:
        return right
    if right is None:
        return left
    x0 = min(left.x, right.x)
    y0 = min(left.y, right.y)
    x1 = max(left.x + left.width, right.x + right.width)
    y1 = max(left.y + left.height, right.y + right.height)
    return BoundingBox(x=x0, y=y0, width=x1 - x0, height=y1 - y0)


def _repeated_pdf_margin_text(
    candidates: list[dict[str, Any]],
    page_count: int,
) -> dict[tuple[int, int], BlockType]:
    if page_count < 2:
        return {}
    occurrences: dict[str, list[dict[str, Any]]] = {}
    for candidate in candidates:
        y0 = float(candidate["bbox"][1])
        y1 = float(candidate["bbox"][3])
        height = max(1.0, float(candidate["pageHeight"]))
        if y0 > height * 0.12 and y1 < height * 0.88:
            continue
        key = re.sub(r"\d+", "#", _clean_inline(str(candidate["text"]))).casefold()
        if len(key) < 2 or len(key) > 180:
            continue
        occurrences.setdefault(key, []).append(candidate)

    threshold = max(2, (page_count + 1) // 2)
    repeated: dict[tuple[int, int], BlockType] = {}
    for values in occurrences.values():
        if len({int(value["page"]) for value in values}) < threshold:
            continue
        for value in values:
            y0 = float(value["bbox"][1])
            kind = (
                BlockType.PAGE_HEADER
                if y0 <= float(value["pageHeight"]) * 0.12
                else BlockType.PAGE_FOOTER
            )
            repeated[(int(value["page"]), int(value["pageBlockIndex"]))] = kind
    return repeated


def _is_list_text(value: str) -> bool:
    lines = [line.strip() for line in value.splitlines() if line.strip()]
    list_lines = sum(1 for line in lines if _is_markdown_list_line(line))
    return bool(lines) and list_lines >= max(1, len(lines) // 2)


_FEATURE_GATE_ENTRY = re.compile(
    r"(?<![\w:])(?:kube:)?[A-Za-z][A-Za-z0-9_.:-]*="
    r"(?:true|false)\|(?:true|false)\s*\(",
    re.IGNORECASE,
)


def _feature_gate_list_text(value: str) -> str | None:
    matches = list(_FEATURE_GATE_ENTRY.finditer(value))
    if not matches:
        return None
    parts: list[str] = []
    prefix = value[: matches[0].start()].strip()
    if prefix:
        parts.append(prefix)
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        item = value[match.start() : end].strip()
        if item:
            parts.append(item)
    return "\n".join(parts)


def _feature_gate_item_count(value: str) -> int:
    return len(_FEATURE_GATE_ENTRY.findall(value))
